"""
Script para regenerar TablaCargaMasiva.xlsx con los nuevos campos del refactor.
Ejecutar desde Backend/: venv/Scripts/python.exe scripts/gen_excel_template.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

NUM_SPECS = 100

wb = openpyxl.Workbook()

# ========== HOJA 1: Productos y Variantes ==========
ws = wb.active
ws.title = 'Productos y Variantes'

headers = [
    'codigo_producto',
    'nombre',
    'marca',
    'categoria_nivel_0',
    'categoria_nivel_1',
    'categoria_nivel_2',
    'categoria_nivel_3',
    'descripcion',
    'nombre_imagen',
    'activo',
    'es_variante',
    'modelo_variante',
    'productos_relacionados',
    'tiene_specs',
    # Columnas posicionales para specs dinamicas (el padre define los nombres, las variantes los valores):
    *[f'spec_{i}' for i in range(1, NUM_SPECS + 1)],
]

header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill('solid', fgColor='1F4E79')
specs_header_fill = PatternFill('solid', fgColor='375623')  # verde oscuro para columnas de specs
variant_fill = PatternFill('solid', fgColor='D6E4F7')

thin = Side(border_style='thin', color='AAAAAA')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
wrap = Alignment(wrap_text=True, vertical='top')

specs_start_col = headers.index('tiene_specs') + 1  # 1-based
is_variant_col_idx = headers.index('es_variante')   # 0-based para row_data

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = specs_header_fill if col_idx >= specs_start_col else header_fill
    cell.alignment = center
    cell.border = border

col_widths = {
    'codigo_producto':      16,
    'nombre':               28,
    'marca':                15,
    'categoria_nivel_0':    17,
    'categoria_nivel_1':    17,
    'categoria_nivel_2':    17,
    'categoria_nivel_3':    17,
    'descripcion':          30,
    'nombre_imagen':        22,
    'activo':               10,
    'es_variante':          12,
    'modelo_variante':      18,
    'productos_relacionados': 22,
    'tiene_specs':          13,
    **{f'spec_{i}': 14 for i in range(1, NUM_SPECS + 1)},
}
for col_idx, h in enumerate(headers, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(h, 14)

_pad = (None,) * (NUM_SPECS - 3)  # relleno para completar hasta NUM_SPECS columnas de specs

example_rows = [
    # Producto 1 — padre define NOMBRES de specs en spec_1/spec_2/... (tiene_specs=true)
    # cols: cod_prod, nombre, marca, cat0, cat1, cat2, cat3, desc, img, activo, es_var, cod_var, relacionados, tiene_specs, spec1, spec2, spec3, ...
    ('MAT-001', 'Matraz Aforado Clase A', 'LabEquip', 'Equipamiento', 'Material Volumetrico', 'Matraces',
     None, 'Matraz aforado de vidrio borosilicato clase A', 'matraz_aforado.jpg', 'true', 'false',
     None, None, 'true', 'potencia', 'velocidad', 'volumen') + _pad,
    # variantes ponen los VALORES en esas mismas columnas
    ('MAT-001', None, None, None, None, None, None, None, None, None, 'true',
     '25ML', None, None, '20hp', '25mhp', '25ml') + _pad,
    ('MAT-001', None, None, None, None, None, None, None, None, None, 'true',
     '50ML', None, None, '40hp', '30mhp', '50ml') + _pad,
    ('MAT-001', None, None, None, None, None, None, None, None, None, 'true',
     '100ML', None, None, '60hp', '40mhp', '100ml') + _pad,
    # Separador
    (None,) * len(headers),
    # Producto 2 — define sus propios nombres de specs (distintos al producto 1)
    ('VAS-002', 'Vaso de Precipitado', 'GlassLab', 'Equipamiento', 'Cristaleria', 'Vasos',
     None, 'Vaso de precipitado graduado', 'vaso_precipitado.jpg', 'true', 'false',
     None, 'MAT-001', 'true', 'viscosidad', 'cap_nominal', None) + _pad,
    ('VAS-002', None, None, None, None, None, None, None, None, None, 'true',
     '100ML', None, None, '52cP', '100ml', None) + _pad,
    ('VAS-002', None, None, None, None, None, None, None, None, None, 'true',
     '250ML', None, None, '48cP', '250ml', None) + _pad,
    ('VAS-002', None, None, None, None, None, None, None, None, None, 'true',
     '500ML', None, None, '45cP', '500ml', None) + _pad,
]

for row_idx, row_data in enumerate(example_rows, 2):
    is_empty = all(v is None for v in row_data)
    is_variant_row = row_data[is_variant_col_idx] == 'true'
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = border
        cell.alignment = wrap
        if not is_empty and is_variant_row:
            cell.fill = variant_fill

ws.freeze_panes = 'A2'
ws.row_dimensions[1].height = 32

# ========== Dropdown para categoria_nivel_0 ==========
cat0_col = get_column_letter(headers.index('categoria_nivel_0') + 1)
dv_cat0 = DataValidation(
    type='list',
    formula1='"Insumos,Procesos,Equipamiento,Mobiliario"',
    allow_blank=True,
    showDropDown=False,  # False = el selector es visible en Excel
    showErrorMessage=True,
    errorTitle='Categoría inválida',
    error='Seleccioná una categoría de la lista: Insumos, Procesos, Equipamiento, Mobiliario.',
    showInputMessage=True,
    promptTitle='Categoría nivel 0',
    prompt='Seleccioná la categoría raíz del producto.',
)
ws.add_data_validation(dv_cat0)
dv_cat0.sqref = f'{cat0_col}2:{cat0_col}5000'

# ========== HOJA 2: Instrucciones ==========
ws2 = wb.create_sheet('Instrucciones')
ws2.column_dimensions['A'].width = 115

title_font = Font(bold=True, size=13, color='1F4E79')
section_font = Font(bold=True, size=11)
normal_font = Font(size=10)
code_font = Font(name='Courier New', size=9, color='8B0000')
title_fill = PatternFill('solid', fgColor='D6E4F7')
code_fill = PatternFill('solid', fgColor='F5F5F5')
error_fill = PatternFill('solid', fgColor='FCE4D6')

instrucciones = [
    ('INSTRUCCIONES PARA CARGA MASIVA DE PRODUCTOS Y VARIANTES', 'title'),
    ('', None),
    ('1. COLUMNAS DISPONIBLES', 'section'),
    ('   codigo_producto        Codigo unico del producto (obligatorio)', 'normal'),
    ('   nombre                 Nombre del producto', 'normal'),
    ('   marca                  Marca del producto', 'normal'),
    ('   categoria_nivel_0      Categoria raiz — debe coincidir con una categoria existente en la base de datos (ej: Insumos, Procesos, Equipamiento, Mobiliario)', 'normal'),
    ('   categoria_nivel_1      Subcategoria nivel 1 (se crea si no existe)', 'normal'),
    ('   categoria_nivel_2      Subcategoria nivel 2 (se crea si no existe; requiere categoria_nivel_1)', 'normal'),
    ('   categoria_nivel_3      Subcategoria nivel 3 (se crea si no existe; requiere categoria_nivel_2)', 'normal'),
    ('   descripcion            Descripcion del producto', 'normal'),
    ('   nombre_imagen          Nombre del archivo de imagen en la libreria (ej: producto.jpg)', 'normal'),
    ('   activo                 true/false — si el producto esta activo (por defecto: true)', 'normal'),
    ('   es_variante            true/false — si la fila es una variante (true) o un producto padre (false)', 'normal'),
    ('   modelo_variante        Modelo unico de la variante (obligatorio para es_variante=true)', 'normal'),
    ('   productos_relacionados Codigos de productos relacionados separados por ; (ej: MAT-001;VAS-002)', 'normal'),
    ('   tiene_specs            true/false — en la fila del producto padre: habilita specs dinamicas para sus variantes', 'normal'),
    ('   spec_1, spec_2, ...    Columnas de specs dinamicas: el padre escribe el NOMBRE, la variante escribe el VALOR', 'normal'),
    ('', None),
    ('2. SPECS DINAMICAS DE VARIANTES', 'section'),
    ('   - Solo aplican a variantes (es_variante=true), nunca al producto padre', 'normal'),
    ('   - En la fila del producto padre, escribir el NOMBRE de cada atributo (ej: potencia, material)', 'normal'),
    ('   - En cada fila de variante, escribir el VALOR correspondiente (ej: 20hp, acero inoxidable)', 'normal'),
    ('   - Columnas sin nombre en la fila del padre se ignoran completamente', 'normal'),
    ('   - Al reimportar una variante existente, sus specs anteriores se reemplazan por las nuevas', 'normal'),
    ('   - El template incluye spec_1 hasta spec_100; podes usar todas las que necesites', 'normal'),
    ('', None),
    ('3. PASOS PARA CARGAR UN PRODUCTO CON VARIANTES', 'section'),
    ('   Paso 1: Una fila con es_variante=false define el producto padre (nombre, marca, categoria, etc.)', 'normal'),
    ('          Si usa specs, poner tiene_specs=true y escribir los nombres de atributos en spec_1, spec_2...', 'normal'),
    ('   Paso 2: Para cada variante, una fila con es_variante=true y el mismo codigo_producto', 'normal'),
    ('          Completar modelo_variante y los valores de specs', 'normal'),
    ('', None),
    ('4. EJEMPLO', 'section'),
    ('   codigo_producto | es_variante | modelo_variante | tiene_specs | spec_1   | spec_2   ', 'code'),
    ('   MAT-001         | false       |                 | true        | potencia | volumen  ', 'code'),
    ('   MAT-001         | true        | 25ML    |             | 20hp     | 25ml     ', 'code'),
    ('   MAT-001         | true        | 50ML    |             | 40hp     | 50ml     ', 'code'),
    ('', None),
    ('5. ERRORES POSIBLES Y SOLUCIONES', 'section'),
    ('', None),
    ('   ERROR: Fila N: variante sin codigo_producto', 'error'),
    ('   Causa:    La fila tiene es_variante=true pero no tiene codigo_producto.', 'normal'),
    ('   Solucion: Completar el campo codigo_producto con el codigo del producto padre.', 'normal'),
    ('', None),
    ('   ERROR: Fila N: codigo_producto "X" duplicado', 'error'),
    ('   Causa:    Hay dos o mas filas con es_variante=false y el mismo codigo_producto.', 'normal'),
    ('   Solucion: Solo puede haber UNA fila de producto por codigo. Las demas deben ser variantes.', 'normal'),
    ('', None),
    ('   ERROR: Fila N: variante sin modelo_variante', 'error'),
    ('   Causa:    La fila tiene es_variante=true pero no tiene modelo_variante.', 'normal'),
    ('   Solucion: Completar el campo modelo_variante con un modelo unico para esa variante.', 'normal'),
    ('', None),
    ('   ERROR: Fila N: producto padre con codigo_producto "X" no encontrado', 'error'),
    ('   Causa:    La variante referencia un producto que no existe en la DB ni en el archivo.', 'normal'),
    ('   Solucion: Agregar la fila del producto padre en el mismo archivo, o verificar el codigo.', 'normal'),
    ('', None),
    ('   NOTA: La columna "modelo_variante" reemplaza a "codigo_variante". Ambas son aceptadas por el importador.', 'normal'),
    ('', None),
    ('   ERROR: Fila N: imagen "X" no encontrada en la libreria de archivos', 'error'),
    ('   Causa:    El nombre_imagen no coincide con ningun archivo en la libreria de imagenes.', 'normal'),
    ('   Solucion: Subir la imagen a la libreria antes de importar, o dejar el campo vacio.', 'normal'),
    ('', None),
    ('   ERROR: Fila N: error al procesar producto "X" — Categoria nivel 0 no encontrada', 'error'),
    ('   Causa:    El valor de categoria_nivel_0 no coincide con ninguna categoria raiz en la base de datos.', 'normal'),
    ('   Solucion: Verificar el nombre exacto de la categoria raiz en el sistema (ej: Insumos, Procesos, Equipamiento, Mobiliario).', 'normal'),
    ('', None),
    ('   ERROR: Fila N: producto relacionado con codigo "X" no encontrado', 'error'),
    ('   Causa:    Un codigo en productos_relacionados no existe en la DB ni en el archivo.', 'normal'),
    ('   Solucion: Verificar el codigo o definir el producto relacionado en el mismo archivo.', 'normal'),
    ('', None),
    ('   NOTA: Si descargaste el Excel de errores y lo corregiste para reimportar,', 'normal'),
    ('         la columna "error" se ignora automaticamente — no hace falta borrarla.', 'normal'),
    ('', None),
    ('6. CAMBIOS DE NOMBRES DE COLUMNAS (version anterior → actual)', 'section'),
    ('   product_code         => codigo_producto', 'normal'),
    ('   name                 => nombre', 'normal'),
    ('   brand                => marca', 'normal'),
    ('   category_level_0/1/2/3 => categoria_nivel_0/1/2/3', 'normal'),
    ('   description          => descripcion', 'normal'),
    ('   image_name           => nombre_imagen', 'normal'),
    ('   is_active            => activo', 'normal'),
    ('   is_variant           => es_variante', 'normal'),
    ('   variant_code         => modelo_variante', 'normal'),
    ('   codigo_variante      => modelo_variante', 'normal'),
    ('   related_product_codes => productos_relacionados', 'normal'),
    ('   is_specs_column      => tiene_specs', 'normal'),
    ('   (ambos formatos son aceptados por el importador)', 'normal'),
]

for row_idx, (text, style) in enumerate(instrucciones, 1):
    cell = ws2.cell(row=row_idx, column=1, value=text)
    if style == 'title':
        cell.font = title_font
        cell.fill = title_fill
    elif style == 'section':
        cell.font = section_font
    elif style == 'error':
        cell.font = Font(bold=True, size=10, color='C00000')
        cell.fill = error_fill
    elif style == 'code':
        cell.font = code_font
        cell.fill = code_fill
    else:
        cell.font = normal_font

wb.save('TablaCargaMasiva.xlsx')
print('TablaCargaMasiva.xlsx actualizado correctamente')
