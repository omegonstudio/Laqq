"""
Script para generar un archivo Excel de ejemplo con variantes.
Ejecutar con: python Backend/products/examples/generar_tabla_ejemplo_variantes.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

# Crear workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Productos y Variantes"

# Headers con formato
headers = [
    'product_code', 'name', 'brand', 'category_level_0', 'category_level_1',
    'category_level_2', 'description', 'image_name', 'is_active',
    'is_variant', 'spec_code', 'volume', 'dimensions', 'cap',
    'outlet', 'accuracy', 'precision', 'additional_specs', 'related_product_codes'
]

# Aplicar estilo a headers
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Ajustar ancho de columnas
column_widths = {
    'A': 15,  # product_code
    'B': 25,  # name
    'C': 15,  # brand
    'D': 18,  # category_level_0
    'E': 20,  # category_level_1
    'F': 20,  # category_level_2
    'G': 40,  # description
    'H': 20,  # image_name
    'I': 12,  # is_active
    'J': 12,  # is_variant
    'K': 20,  # spec_code
    'L': 12,  # volume
    'M': 15,  # dimensions
    'N': 12,  # cap
    'O': 12,  # outlet
    'P': 12,  # accuracy
    'Q': 12,  # precision
    'R': 30,  # additional_specs
    'S': 25,  # related_product_codes
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Ejemplo 1: Producto con múltiples variantes
variant_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

# Producto principal
ws.append([
    'MAT-001',                          # product_code
    'Matraz Aforado Clase A',           # name
    'LabEquip',                         # brand
    'equipos',                          # category_level_0
    'Material Volumétrico',             # category_level_1
    'Matraces',                         # category_level_2
    'Matraz aforado de vidrio borosilicato, clase A, con certificado',  # description
    'matraz_aforado.jpg',               # image_name
    'true',                             # is_active
    'false',                            # is_variant (ES UN PRODUCTO)
    'MAT-001-25ML',                     # spec_code
    '25ml',                             # volume
    '70x40mm',                          # dimensions
    'NS 10/19',                         # cap
    '',                                 # outlet
    '±0.03ml',                          # accuracy
    '0.01ml',                           # precision
    '',                                 # additional_specs
    '',                                 # related_product_codes
])

# Variantes del producto (con color de fondo diferente)
variantes_mat001 = [
    ['MAT-001', '', '', '', '', '', '', '', '', 'true', 'MAT-001-50ML', '50ml', '90x45mm', 'NS 12/21', '', '±0.05ml', '0.02ml', '', ''],
    ['MAT-001', '', '', '', '', '', '', '', '', 'true', 'MAT-001-100ML', '100ml', '110x50mm', 'NS 12/21', '', '±0.08ml', '0.02ml', '', ''],
    ['MAT-001', '', '', '', '', '', '', '', '', 'true', 'MAT-001-250ML', '250ml', '140x60mm', 'NS 14/23', '', '±0.15ml', '0.05ml', '', ''],
]

for variant_row in variantes_mat001:
    row_num = ws.max_row + 1
    ws.append(variant_row)
    # Aplicar color de fondo a las variantes
    for col_num in range(1, len(headers) + 1):
        ws.cell(row=row_num, column=col_num).fill = variant_fill

# Ejemplo 2: Otro producto con variantes
ws.append([''])  # Fila vacía para separar

ws.append([
    'VAS-002',                          # product_code
    'Vaso de Precipitado',              # name
    'GlassLab',                         # brand
    'equipos',                          # category_level_0
    'Cristalería',                      # category_level_1
    'Vasos',                            # category_level_2
    'Vaso de precipitado graduado',     # description
    'vaso_precipitado.jpg',             # image_name
    'true',                             # is_active
    'false',                            # is_variant (ES UN PRODUCTO)
    'VAS-002-100ML',                    # spec_code
    '100ml',                            # volume
    '50x70mm',                          # dimensions
    'Sin tapa',                         # cap
    'Pico vertedor',                    # outlet
    '',                                 # accuracy
    '',                                 # precision
    '{"material": "vidrio borosilicato"}',  # additional_specs
    '',                                 # related_product_codes
])

variantes_vas002 = [
    ['VAS-002', '', '', '', '', '', '', '', '', 'true', 'VAS-002-250ML', '250ml', '70x95mm', 'Sin tapa', 'Pico vertedor', '', '', '{"material": "vidrio borosilicato"}', ''],
    ['VAS-002', '', '', '', '', '', '', '', '', 'true', 'VAS-002-500ML', '500ml', '90x125mm', 'Sin tapa', 'Pico vertedor', '', '', '{"material": "vidrio borosilicato"}', ''],
    ['VAS-002', '', '', '', '', '', '', '', '', 'true', 'VAS-002-1000ML', '1000ml', '110x155mm', 'Sin tapa', 'Pico vertedor', '', '', '{"material": "vidrio borosilicato"}', ''],
]

for variant_row in variantes_vas002:
    row_num = ws.max_row + 1
    ws.append(variant_row)
    for col_num in range(1, len(headers) + 1):
        ws.cell(row=row_num, column=col_num).fill = variant_fill

# Ejemplo 3: Producto simple sin variantes
ws.append([''])  # Fila vacía

ws.append([
    'TERM-003',                         # product_code
    'Termómetro Digital',               # name
    'TempTech',                         # brand
    'equipos',                          # category_level_0
    'Medición',                         # category_level_1
    'Temperatura',                      # category_level_2
    'Termómetro digital de alta precisión',  # description
    'termometro.jpg',                   # image_name
    'true',                             # is_active
    'false',                            # is_variant (ES UN PRODUCTO SIN VARIANTES)
    '',                                 # spec_code (vacío, no tiene variantes)
    '',                                 # volume
    '',                                 # dimensions
    '',                                 # cap
    '',                                 # outlet
    '±0.1°C',                           # accuracy
    '0.1°C',                            # precision
    '{"rango": "-50°C a 300°C", "bateria": "AAA"}',  # additional_specs
    '',                                 # related_product_codes
])

# Crear hoja de instrucciones
ws_instrucciones = wb.create_sheet("Instrucciones")
ws_instrucciones.column_dimensions['A'].width = 100

instrucciones = [
    "INSTRUCCIONES PARA CARGA MASIVA CON VARIANTES",
    "",
    "1. COLUMNAS IMPORTANTES:",
    "   - product_code: Código único del producto (obligatorio)",
    "   - is_variant: Indica si es una variante (true) o un producto (false)",
    "   - spec_code: Código único de la variante (obligatorio solo para variantes)",
    "",
    "2. CREAR UN PRODUCTO CON VARIANTES:",
    "   Paso 1: Crear el producto principal con is_variant=false",
    "   Paso 2: Agregar variantes con is_variant=true y el MISMO product_code",
    "",
    "3. REGLAS:",
    "   - Solo puede haber UN producto (is_variant=false) por product_code",
    "   - Las variantes (is_variant=true) DEBEN tener un producto padre existente",
    "   - Las variantes pueden dejar vacíos los campos del producto (name, brand, etc.)",
    "   - Cada variante debe tener un spec_code único",
    "",
    "4. EJEMPLO:",
    "   Fila 1: MAT-001, 'Matraz', 'Marca', 'equipos', ..., is_variant=false, spec_code=MAT-001-25ML",
    "   Fila 2: MAT-001, '', '', '', ..., is_variant=true, spec_code=MAT-001-50ML",
    "   Fila 3: MAT-001, '', '', '', ..., is_variant=true, spec_code=MAT-001-100ML",
    "",
    "5. ERRORES COMUNES:",
    "   ❌ Dos productos con el mismo product_code y is_variant=false",
    "   ❌ Variante (is_variant=true) sin producto padre",
    "   ❌ Variante sin spec_code",
    "",
    "6. CATEGORÍAS:",
    "   - category_level_0: Solo puede ser: insumos, procesos, equipos, mobiliario",
    "   - category_level_1 y level_2: Se crean automáticamente si no existen",
    "",
    "7. Ver documentación completa en: Backend/products/README_CARGA_MASIVA_VARIANTES.md",
]

for i, instruccion in enumerate(instrucciones, 1):
    cell = ws_instrucciones.cell(row=i, column=1)
    cell.value = instruccion
    if i == 1:
        cell.font = Font(size=14, bold=True)
    cell.alignment = Alignment(wrap_text=True, vertical='top')

# Guardar archivo (sobreescribir el archivo existente)
output_path = os.path.join(os.path.dirname(__file__), '..', '..', 'TablaCargaMasiva.xlsx')
wb.save(output_path)
print(f"OK - Archivo actualizado exitosamente: {output_path}")
print("El archivo incluye:")
print("   - Hoja 'Productos y Variantes': Ejemplos de carga")
print("   - Hoja 'Instrucciones': Guia de uso")
