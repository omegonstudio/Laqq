from django.test import TestCase
from rest_framework.test import APITestCase, APIClient
from rest_framework import status
from django.contrib.auth import get_user_model
from .models import Brand, Category, Product, ProductVariant, ProductRelation, TechnicalSpec, VariantTechnicalSpec
from attachments.models import Attachment
from users.models import UserType

User = get_user_model()


class BrandAPITestCase(APITestCase):
    """Tests para el CRUD de Marcas de productos"""

    def setUp(self):
        self.client = APIClient()
        self.user_type = UserType.objects.create(id='admin', name='Admin')
        self.user = User.objects.create_user(username='testuser', password='testpass123')
        self.user.user_type = self.user_type
        self.user.save()
        self.client.force_authenticate(user=self.user)
        self.brand = Brand.objects.create(name='Test Brand', description='Test Description')

    def test_list_brands(self):
        """Listar todas las marcas con paginación"""
        response = self.client.get('/products/brands/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 1)

    def test_create_brand(self):
        """Crear una nueva marca con nombre y descripción"""
        data = {'name': 'New Brand', 'description': 'New Description'}
        response = self.client.post('/products/brands/', data)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(Brand.objects.count(), 2)

    def test_update_brand(self):
        """Actualizar nombre y descripción de una marca existente"""
        data = {'name': 'Updated Brand', 'description': 'Updated Description'}
        response = self.client.put(f'/products/brands/{self.brand.id}/', data)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.brand.refresh_from_db()
        self.assertEqual(self.brand.name, 'Updated Brand')

    def test_delete_brand(self):
        """Eliminar una marca del sistema"""
        response = self.client.delete(f'/products/brands/{self.brand.id}/')
        self.assertEqual(response.status_code, status.HTTP_204_NO_CONTENT)
        self.assertEqual(Brand.objects.count(), 0)

    def test_search_brand(self):
        """Buscar marcas por nombre usando el parámetro search"""
        Brand.objects.create(name='Another Brand')
        response = self.client.get('/products/brands/?search=Test')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 1)

    def test_brand_with_logo_returns_url(self):
        """Verificar que logo_url se devuelve cuando hay logo_attachment"""
        from django.core.files.uploadedfile import SimpleUploadedFile

        test_file = SimpleUploadedFile(
            'brand-logo.png',
            b'fake image content',
            content_type='image/png'
        )

        attachment = Attachment.objects.create(
            file=test_file,
            role='image',
            content_type_str='image/png',
            attachable_type='brand',
            attachable_id=self.brand.id
        )

        self.brand.logo_attachment = attachment
        self.brand.save()

        response = self.client.get(f'/products/brands/{self.brand.id}/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIsNotNone(response.data['logo_url'])
        self.assertIn('logo_url', response.data)
        self.assertTrue(isinstance(response.data['logo_url'], str))

    def test_delete_brand_sets_products_brand_to_null(self):
        """Verificar que al borrar una marca, los productos quedan con brand=None"""
        category = Category.objects.create(name='Test Category')
        product = Product.objects.create(
            name='Product with Brand',
            brand=self.brand,
            category=category
        )

        self.brand.delete()

        product.refresh_from_db()
        self.assertIsNone(product.brand)
        self.assertEqual(Product.objects.count(), 1)


class CategoryAPITestCase(APITestCase):
    """Tests para el CRUD de Categorías de productos"""

    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(username='testuser', password='testpass123')
        self.client.force_authenticate(user=self.user)
        self.category = Category.objects.create(name='Test Category', display_order=1)

    def test_list_categories(self):
        """Listar todas las categorías con paginación"""
        response = self.client.get('/products/categories/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 1)

    def test_create_category(self):
        """Crear una nueva categoría con nombre y orden de visualización"""
        data = {'name': 'New Category', 'display_order': 2}
        response = self.client.post('/products/categories/', data)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(Category.objects.count(), 2)

    def test_create_category_with_level(self):
        """Crear categoría con nivel específico (usando parent para auto-calcular)"""
        parent = Category.objects.create(name='Parent Category', display_order=1)

        data = {
            'name': 'Subcategory Level 1',
            'parent': parent.id,
            'display_order': 2
        }
        response = self.client.post('/products/categories/', data)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data['level'], 1)
        category = Category.objects.get(name='Subcategory Level 1')
        self.assertEqual(category.level, 1)

    def test_category_hierarchy_levels(self):
        """Verificar niveles en jerarquía de categorías"""
        parent = Category.objects.create(name='Parent Category', level=0)
        child = Category.objects.create(name='Child Category', parent=parent, level=1)
        grandchild = Category.objects.create(name='Grandchild Category', parent=child, level=2)

        response = self.client.get(f'/products/categories/{parent.id}/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['level'], 0)
        self.assertIsNone(response.data['parent'])

        response = self.client.get(f'/products/categories/{child.id}/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['level'], 1)
        self.assertEqual(str(response.data['parent']), str(parent.id))

        response = self.client.get(f'/products/categories/{grandchild.id}/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['level'], 2)
        self.assertEqual(str(response.data['parent']), str(child.id))

    def test_auto_calculate_level_without_parent(self):
        """Verificar que level se auto-calcula como 0 cuando no hay parent"""
        category = Category.objects.create(name='Root Category', display_order=1)
        self.assertEqual(category.level, 0)

    def test_auto_calculate_level_with_parent(self):
        """Verificar que level se auto-calcula basándose en el parent"""
        parent = Category.objects.create(name='Parent', display_order=1)
        self.assertEqual(parent.level, 0)

        child = Category.objects.create(name='Child', parent=parent, display_order=2)
        self.assertEqual(child.level, 1)

        grandchild = Category.objects.create(name='Grandchild', parent=child, display_order=3)
        self.assertEqual(grandchild.level, 2)

    def test_auto_calculate_level_via_api(self):
        """Verificar que level se auto-calcula al crear categoría via API"""
        parent_data = {'name': 'API Parent', 'display_order': 1}
        parent_response = self.client.post('/products/categories/', parent_data)
        self.assertEqual(parent_response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(parent_response.data['level'], 0)
        parent_id = parent_response.data['id']

        child_data = {
            'name': 'API Child',
            'parent': parent_id,
            'display_order': 2
        }
        child_response = self.client.post('/products/categories/', child_data)
        self.assertEqual(child_response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(child_response.data['level'], 1)

    def test_delete_category_sets_products_category_to_null(self):
        """Verificar que al borrar una categoría, los productos quedan con category=None"""
        brand = Brand.objects.create(name='Test Brand')
        product = Product.objects.create(
            name='Product with Category',
            brand=brand,
            category=self.category
        )

        self.category.delete()

        product.refresh_from_db()
        self.assertIsNone(product.category)
        self.assertEqual(Product.objects.count(), 1)


class ProductAPITestCase(APITestCase):
    """Tests para el CRUD de Productos"""

    def setUp(self):
        self.client = APIClient()
        self.user_type = UserType.objects.create(id='admin', name='Admin')
        self.user = User.objects.create_user(username='testuser', password='testpass123', is_staff=True)
        self.user.user_type = self.user_type
        self.user.save()
        self.client.force_authenticate(user=self.user)
        self.brand = Brand.objects.create(name='Test Brand')
        self.category = Category.objects.create(name='Test Category')
        self.product = Product.objects.create(
            name='Test Product',
            brand=self.brand,
            category=self.category,
            is_active=True
        )

    def test_list_products(self):
        """Listar todos los productos con paginación"""
        response = self.client.get('/products/list/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 1)
        self.assertIn('count', response.data)
        self.assertIn('next', response.data)
        self.assertIn('previous', response.data)
        self.assertIn('page_size', response.data)
        self.assertIn('current_page', response.data)
        self.assertIn('total_pages', response.data)
        self.assertEqual(response.data['current_page'], 1)
        self.assertEqual(response.data['total_pages'], 1)
        self.assertEqual(response.data['page_size'], 25)

    def test_create_product(self):
        """Crear un nuevo producto asociado a marca y categoría"""
        data = {
            'name': 'New Product',
            'brand_id': self.brand.id,
            'category_id': self.category.id,
            'is_active': True
        }
        response = self.client.post('/products/list/', data)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(Product.objects.count(), 2)
        self.assertEqual(response.data['brand'], self.brand.name)
        self.assertEqual(response.data['category'], self.category.name)

    def test_update_spec_table(self):
        """Se puede guardar una tabla de especificaciones con filas y columnas."""
        payload = {
            'spec_table': {
                'columns': ['Parámetro', 'Valor'],
                'rows': [['Voltaje', '220V'], ['Frecuencia', '50Hz']],
            }
        }
        response = self.client.patch(
            f'/products/list/{self.product.id}/',
            payload,
            format='json',
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['spec_table']['columns'], ['Parámetro', 'Valor'])
        self.assertEqual(len(response.data['spec_table']['rows']), 2)
        self.product.refresh_from_db()
        self.assertEqual(self.product.spec_table['rows'][0][1], '220V')

    def test_invalid_spec_table_rejected(self):
        response = self.client.patch(
            f'/products/list/{self.product.id}/',
            {'spec_table': ['no', 'es', 'objeto']},
            format='json',
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_filter_by_brand(self):
        """Filtrar productos por marca específica"""
        response = self.client.get(f'/products/list/?brand={self.brand.id}')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 1)

    def test_filter_by_active(self):
        """Filtrar productos por estado activo/inactivo"""
        Product.objects.create(
            name='Inactive Product',
            brand=self.brand,
            category=self.category,
            is_active=False
        )
        response = self.client.get('/products/list/?is_active=true')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 1)

    def test_search_product(self):
        """Buscar productos por nombre usando el parámetro search"""
        response = self.client.get('/products/list/?search=Test')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 1)

    def test_filter_featured_products(self):
        """Filtrar solo productos destacados"""
        Product.objects.create(
            name='Featured Product',
            brand=self.brand,
            category=self.category,
            is_active=True,
            is_featured=True
        )
        Product.objects.create(
            name='Regular Product',
            brand=self.brand,
            category=self.category,
            is_active=True,
            is_featured=False
        )
        response = self.client.get('/products/list/?is_featured=true')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 1)
        self.assertEqual(response.data['results'][0]['name'], 'Featured Product')


class ProductVariantAPITestCase(APITestCase):
    """Tests para el CRUD de Variantes de productos"""

    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(username='testuser', password='testpass123')
        self.client.force_authenticate(user=self.user)
        self.brand = Brand.objects.create(name='Test Brand')
        self.category = Category.objects.create(name='Test Category')
        self.product = Product.objects.create(
            name='Test Product',
            brand=self.brand,
            category=self.category
        )
        self.variant = ProductVariant.objects.create(
            product=self.product,
            code='TEST-001',
            name='Variante A'
        )

    def test_list_variants(self):
        """Listar todas las variantes de productos"""
        response = self.client.get('/products/variants/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 1)

    def test_create_variant(self):
        """Crear una nueva variante con código y nombre"""
        data = {
            'product': self.product.id,
            'code': 'TEST-002',
            'name': 'Variante B'
        }
        response = self.client.post('/products/variants/', data)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(ProductVariant.objects.count(), 2)

    def test_filter_by_product(self):
        """Filtrar variantes por producto específico"""
        response = self.client.get(f'/products/variants/?product={self.product.id}')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['results']), 1)


class BulkUploadBackofficeTestCase(APITestCase):
    """Un usuario BACKOFFICE puede subir productos vía Excel y el producto se limpia al final"""

    def setUp(self):
        import openpyxl
        from io import BytesIO
        from django.core.files.uploadedfile import SimpleUploadedFile

        self.client = APIClient()

        backoffice_type, _ = UserType.objects.get_or_create(id='BACKOFFICE', defaults={'name': 'Backoffice'})
        self.user = User.objects.create_user(username='backoffice_test', password='testpass123')
        self.user.user_type = backoffice_type
        self.user.save()
        self.client.force_authenticate(user=self.user)

        # categoria_nivel_0 debe pre-existir
        self.root_category = Category.objects.create(name='Consumibles Test', display_order=99)

        # Generar Excel en memoria con una fila de producto
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = [
            'codigo_producto', 'nombre', 'marca',
            'categoria_nivel_0', 'categoria_nivel_1',
            'descripcion', 'activo', 'es_variante',
        ]
        ws.append(headers)
        ws.append([
            'TUBO-PRUEBA-001', 'tubo de prueba', 'Marca Test',
            'Consumibles Test', 'Tubos',
            'Producto de prueba para test automatizado', '1', '0',
        ])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        self.excel_file = SimpleUploadedFile(
            'productos_test.xlsx',
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_backoffice_puede_hacer_carga_masiva(self):
        """BACKOFFICE sube un Excel con 'tubo de prueba', verifica creación y luego lo borra"""
        response = self.client.post(
            '/products/bulk-upload/',
            {'csv_file': self.excel_file},
            format='multipart',
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data.get('created_products'), 1)
        self.assertEqual(len(response.data.get('errors', [])), 0)

        product = Product.objects.filter(name='tubo de prueba').first()
        self.assertIsNotNone(product, 'El producto "tubo de prueba" debería existir en la base de datos')
        self.assertEqual(product.product_code, 'TUBO-PRUEBA-001')

        # Limpieza explícita
        product.delete()
        self.assertFalse(Product.objects.filter(name='tubo de prueba').exists())


class BulkUploadInsumoTestCase(APITestCase):
    """Carga masiva de Consumibles: campos articulo/cas/sedronar + PDFs ESP/HDS por nombre."""

    def setUp(self):
        import openpyxl
        from io import BytesIO
        from django.core.files.uploadedfile import SimpleUploadedFile

        self.client = APIClient()

        backoffice_type, _ = UserType.objects.get_or_create(id='BACKOFFICE', defaults={'name': 'Backoffice'})
        self.user = User.objects.create_user(username='backoffice_consumible', password='testpass123')
        self.user.user_type = backoffice_type
        self.user.save()
        self.client.force_authenticate(user=self.user)

        # categoria_nivel_0 debe pre-existir
        self.root_category = Category.objects.create(name='Consumibles Test', display_order=99)
        self.subcategory = Category.objects.create(name='Reactivos', parent=self.root_category)

        # Attachments existentes en la librería (los PDFs que el Excel va a referenciar por nombre)
        esp_file = SimpleUploadedFile('reactivo-x-esp.pdf', b'%PDF-1.4 fake esp', content_type='application/pdf')
        hds_file = SimpleUploadedFile('reactivo-x-hds.pdf', b'%PDF-1.4 fake hds', content_type='application/pdf')
        self.esp_attachment = Attachment.objects.create(
            file=esp_file, role='other', content_type_str='application/pdf',
            attachable_type='library', attachable_id=None,
        )
        self.hds_attachment = Attachment.objects.create(
            file=hds_file, role='other', content_type_str='application/pdf',
            attachable_type='library', attachable_id=None,
        )

    def _build_excel(self, rows):
        """Helper: arma un Excel con las columnas de insumo + 1 fila (o más)."""
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = [
            'codigo_producto', 'nombre', 'marca',
            'categoria_nivel_0', 'categoria_nivel_1',
            'descripcion', 'activo', 'es_variante',
            'articulo', 'cas', 'sedronar', 'archivo_esp', 'archivo_hds',
        ]
        ws.append(headers)
        for r in rows:
            ws.append(r)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return SimpleUploadedFile(
            'consumibles_test.xlsx',
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_importa_insumo_con_campos_y_pdfs(self):
        """Fila de insumo con articulo/cas/sedronar + ambos PDFs por nombre → todo se persiste."""
        excel = self._build_excel([[
            'INS-001', 'Reactivo X', 'Marca Lab',
            'Consumibles Test', 'Reactivos',
            'Descripción de prueba', '1', '0',
            'ART-123', '7732-18-5', 'SED-001',
            'reactivo-x-esp.pdf', 'reactivo-x-hds.pdf',
        ]])
        response = self.client.post(
            '/products/bulk-upload/', {'csv_file': excel}, format='multipart',
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data.get('created_products'), 1)
        self.assertEqual(response.data.get('linked_esp_pdfs'), 1)
        self.assertEqual(response.data.get('linked_hds_pdfs'), 1)

        product = Product.objects.get(product_code='INS-001')
        self.assertEqual(product.articulo, 'ART-123')
        self.assertEqual(product.cas, '7732-18-5')
        self.assertEqual(product.sedronar, 'SED-001')
        # root_category se calcula en save() desde la jerarquía
        self.assertEqual(product.root_category, 'consumibles test')
        self.assertEqual(product.esp_attachment_id, self.esp_attachment.id)
        self.assertEqual(product.hds_attachment_id, self.hds_attachment.id)

    def test_pdf_inexistente_genera_warning_sin_abortar(self):
        """Si el PDF referenciado no existe en la librería, se crea el producto igual
        y se agrega una entrada en summary['errors'] con la fila."""
        excel = self._build_excel([[
            'INS-002', 'Reactivo Y', 'Marca Lab',
            'Consumibles Test', 'Reactivos',
            '', '1', '0',
            'ART-999', '', '',
            'no-existe-esp.pdf', '',
        ]])
        response = self.client.post(
            '/products/bulk-upload/', {'csv_file': excel}, format='multipart',
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data.get('created_products'), 1)

        errors = response.data.get('errors', [])
        self.assertEqual(len(errors), 1)
        self.assertIn('no-existe-esp.pdf', errors[0]['error'])
        self.assertEqual(errors[0]['row'], 2)

        # El producto se creó igual, sin ESP pero con sus textos
        product = Product.objects.get(product_code='INS-002')
        self.assertEqual(product.articulo, 'ART-999')
        self.assertIsNone(product.esp_attachment)

    def test_reimportar_con_campos_vacios_preserva_valores(self):
        """Reimportación con celdas vacías en campos de insumo NO pisa los valores previos."""
        # 1ra importación: completa
        excel1 = self._build_excel([[
            'INS-003', 'Reactivo Z', 'Marca Lab',
            'Consumibles Test', 'Reactivos',
            '', '1', '0',
            'ART-ORIG', '111-11-1', 'SED-ORIG',
            'reactivo-x-esp.pdf', 'reactivo-x-hds.pdf',
        ]])
        self.client.post('/products/bulk-upload/', {'csv_file': excel1}, format='multipart')
        product = Product.objects.get(product_code='INS-003')
        self.assertEqual(product.articulo, 'ART-ORIG')
        self.assertEqual(product.cas, '111-11-1')
        self.assertEqual(product.sedronar, 'SED-ORIG')
        original_esp_id = product.esp_attachment_id
        original_hds_id = product.hds_attachment_id

        # 2da importación: celdas de insumo y PDFs vacías → debe preservar todo
        excel2 = self._build_excel([[
            'INS-003', 'Reactivo Z', 'Marca Lab',
            'Consumibles Test', 'Reactivos',
            '', '1', '0',
            '', '', '',
            '', '',
        ]])
        response = self.client.post('/products/bulk-upload/', {'csv_file': excel2}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data.get('updated_products'), 1)
        self.assertEqual(response.data.get('linked_esp_pdfs'), 0)
        self.assertEqual(response.data.get('linked_hds_pdfs'), 0)

        product.refresh_from_db()
        self.assertEqual(product.articulo, 'ART-ORIG', 'articulo no debe sobrescribirse con vacío')
        self.assertEqual(product.cas, '111-11-1', 'cas no debe sobrescribirse con vacío')
        self.assertEqual(product.sedronar, 'SED-ORIG', 'sedronar no debe sobrescribirse con vacío')
        self.assertEqual(product.esp_attachment_id, original_esp_id, 'esp_attachment no debe limpiarse con celda vacía')
        self.assertEqual(product.hds_attachment_id, original_hds_id, 'hds_attachment no debe limpiarse con celda vacía')

    def test_producto_no_insumo_ignora_campos_vacios(self):
        """Si el producto no es insumo (otra categoría), las columnas vienen vacías sin error."""
        other_root = Category.objects.create(name='Equipos Test', display_order=98)
        other_sub = Category.objects.create(name='Microscopios', parent=other_root)
        excel = self._build_excel([[
            'EQ-001', 'Microscopio Óptico', 'Marca Lab',
            'Equipos Test', 'Microscopios',
            '', '1', '0',
            '', '', '',
            '', '',
        ]])
        response = self.client.post(
            '/products/bulk-upload/', {'csv_file': excel}, format='multipart',
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data.get('created_products'), 1)
        self.assertEqual(len(response.data.get('errors', [])), 0)

        product = Product.objects.get(product_code='EQ-001')
        self.assertEqual(product.root_category, 'equipos test')
        self.assertEqual(product.articulo, '')
        self.assertEqual(product.cas, '')
        self.assertEqual(product.sedronar, '-')  # default del modelo
        self.assertIsNone(product.esp_attachment)
        self.assertIsNone(product.hds_attachment)


class ProductExportTestCase(APITestCase):
    """Excel de productos en formato CargaMasiva."""

    def setUp(self):
        self.client = APIClient()
        self.user_type = UserType.objects.create(id='admin', name='Admin')
        self.user = User.objects.create_user(username='exportadmin', password='testpass123')
        self.user.user_type = self.user_type
        self.user.save()
        self.client.force_authenticate(user=self.user)

        self.root = Category.objects.create(name='Equipos')
        self.sub = Category.objects.create(name='Centrifugas', parent=self.root)
        self.brand_a = Brand.objects.create(name='LabEquip')
        self.brand_b = Brand.objects.create(name='OtraMarca')

        self.active = Product.objects.create(
            name='Centrifuga Activa',
            product_code='CEN-001',
            brand=self.brand_a,
            category=self.sub,
            is_active=True,
            description='Equipo de lab',
        )
        self.inactive = Product.objects.create(
            name='Centrifuga Inactiva',
            product_code='CEN-002',
            brand=self.brand_b,
            category=self.sub,
            is_active=False,
        )
        self.consumible_root = Category.objects.create(name='Consumibles')
        self.consumible = Product.objects.create(
            name='Acido clorhidrico',
            product_code='CON-001',
            brand=self.brand_a,
            category=self.consumible_root,
            is_active=True,
            articulo='ART-123',
            cas='7647-01-0',
            sedronar='Lista 1',
        )
        ProductRelation.objects.create(from_product=self.consumible, to_product=self.active)

        v1 = ProductVariant.objects.create(product=self.active, code='CEN-001-A')
        v2 = ProductVariant.objects.create(product=self.active, code='CEN-001-B')
        spec_a = TechnicalSpec.objects.create(key='potencia', value='20hp')
        spec_b = TechnicalSpec.objects.create(key='potencia', value='40hp')
        VariantTechnicalSpec.objects.create(variant=v1, technical_spec=spec_a)
        VariantTechnicalSpec.objects.create(variant=v2, technical_spec=spec_b)

    def _load_sheet(self, response):
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(response.content))
        return wb.active

    def test_anonymous_cannot_export(self):
        self.client.force_authenticate(user=None)
        response = self.client.get('/products/list/export/')
        self.assertIn(response.status_code, (status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN))

    def test_export_includes_inactive_and_variants(self):
        response = self.client.get('/products/list/export/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn(
            'spreadsheetml',
            response['Content-Type'],
        )
        ws = self._load_sheet(response)
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers[0], 'codigo_producto')
        self.assertIn('articulo', headers)
        self.assertIn('es_variante', headers)
        self.assertIn('spec_1', headers)
        self.assertNotIn('spec_table', headers)

        rows = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2, values_only=False)]
        codes = [row[0] for row in rows]
        self.assertIn('CEN-001', codes)
        self.assertIn('CEN-002', codes)
        self.assertIn('CON-001', codes)

        parent = next(row for row in rows if row[0] == 'CEN-001' and row[headers.index('es_variante')] == 'false')
        self.assertEqual(parent[headers.index('categoria_nivel_0')], 'Equipos')
        self.assertEqual(parent[headers.index('categoria_nivel_1')], 'Centrifugas')
        self.assertEqual(parent[headers.index('tiene_specs')], 'true')
        self.assertEqual(parent[headers.index('spec_1')], 'potencia')

        variant_codes = [
            row[headers.index('modelo_variante')]
            for row in rows
            if row[0] == 'CEN-001' and row[headers.index('es_variante')] == 'true'
        ]
        self.assertEqual(sorted(variant_codes), ['CEN-001-A', 'CEN-001-B'])

        consumible = next(row for row in rows if row[0] == 'CON-001')
        self.assertEqual(consumible[headers.index('articulo')], 'ART-123')
        self.assertEqual(consumible[headers.index('cas')], '7647-01-0')
        self.assertEqual(consumible[headers.index('productos_relacionados')], 'CEN-001')
        self.assertEqual(consumible[headers.index('activo')], 'true')

        inactive = next(row for row in rows if row[0] == 'CEN-002')
        self.assertEqual(inactive[headers.index('activo')], 'false')

    def test_export_applies_brand_filter(self):
        response = self.client.get(f'/products/list/export/?brand={self.brand_b.id}')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        ws = self._load_sheet(response)
        codes = {row[0].value for row in ws.iter_rows(min_row=2) if row[0].value}
        self.assertEqual(codes, {'CEN-002'})

    def test_export_applies_search(self):
        response = self.client.get('/products/list/export/?search=Acido')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        ws = self._load_sheet(response)
        codes = {row[0].value for row in ws.iter_rows(min_row=2) if row[0].value}
        self.assertEqual(codes, {'CON-001'})

