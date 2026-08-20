"""Exporta productos al mismo formato Excel que la carga masiva."""
from io import BytesIO

from django.db.models import Prefetch

from .models import Product, ProductVariant

EXCEL_MAX_CELL = 32767

FIXED_HEADERS = [
    'codigo_producto',
    'nombre',
    'marca',
    'categoria_nivel_0',
    'categoria_nivel_1',
    'categoria_nivel_2',
    'categoria_nivel_3',
    'descripcion',
    'nombre_imagen',
    'articulo',
    'cas',
    'sedronar',
    'archivo_esp',
    'archivo_hds',
    'activo',
    'es_variante',
    'modelo_variante',
    'productos_relacionados',
    'tiene_specs',
]


def export_queryset():
    """Queryset con joins necesarios para no N+1 al armar el Excel."""
    return (
        Product.objects.select_related(
            'brand',
            'category',
            'category__parent',
            'category__parent__parent',
            'category__parent__parent__parent',
            'image_attachment',
            'esp_attachment',
            'hds_attachment',
        )
        .prefetch_related(
            Prefetch(
                'variants',
                queryset=ProductVariant.objects.order_by('created_at').prefetch_related(
                    'technical_specs'
                ),
            ),
            'from_relations__to_product',
        )
        .order_by('-created_at')
    )


def _cell(value):
    if value is None:
        return ''
    text = str(value)
    if len(text) > EXCEL_MAX_CELL:
        return text[:EXCEL_MAX_CELL]
    return text


def _bool_cell(value):
    return 'true' if value else 'false'


def _attachment_name(attachment):
    if not attachment:
        return ''
    return attachment.file_name or ''


def _category_levels(product):
    names = []
    seen = set()
    current = product.category
    while current is not None and current.id not in seen:
        seen.add(current.id)
        names.append(current.name)
        current = current.parent
    names.reverse()
    names.extend([''] * (4 - len(names)))
    return names[:4]


def _related_codes(product):
    codes = []
    for relation in product.from_relations.all():
        code = getattr(relation.to_product, 'product_code', '') or ''
        if code:
            codes.append(code)
    return ';'.join(codes)


def _spec_keys(product):
    keys = []
    seen = set()
    for variant in product.variants.all():
        for spec in variant.technical_specs.all():
            key = (spec.key or '').strip()
            if key and key not in seen:
                seen.add(key)
                keys.append(key)
    return keys


def _spec_map(variant):
    return {
        (spec.key or '').strip(): spec.value or ''
        for spec in variant.technical_specs.all()
        if (spec.key or '').strip()
    }


def build_products_workbook(products):
    """
    Arma un .xlsx con las mismas columnas que TablaCargaMasiva.xlsx.
    `spec_1`…`spec_N` solo hasta el máximo de keys de un producto.
    No incluye spec_table.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    products = list(products)
    spec_keys_by_id = {product.id: _spec_keys(product) for product in products}
    max_specs = max((len(keys) for keys in spec_keys_by_id.values()), default=0)
    spec_headers = [f'spec_{i}' for i in range(1, max_specs + 1)]
    headers = FIXED_HEADERS + spec_headers

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Productos y Variantes'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='4472C4')
    thin = Side(border_style='thin', color='AAAAAA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wrap = Alignment(wrap_text=True, vertical='top')

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for product in products:
        keys = spec_keys_by_id[product.id]
        levels = _category_levels(product)
        padded_keys = keys + [''] * (max_specs - len(keys))
        parent_row = [
            product.product_code or '',
            product.name or '',
            getattr(product.brand, 'name', '') or '',
            levels[0],
            levels[1],
            levels[2],
            levels[3],
            product.description or '',
            _attachment_name(product.image_attachment),
            product.articulo or '',
            product.cas or '',
            product.sedronar or '',
            _attachment_name(product.esp_attachment),
            _attachment_name(product.hds_attachment),
            _bool_cell(product.is_active),
            'false',
            '',
            _related_codes(product),
            _bool_cell(bool(keys)),
            *padded_keys,
        ]
        row_idx = ws.max_row + 1
        for col_idx, value in enumerate(parent_row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_cell(value))
            cell.alignment = wrap
            cell.border = border

        for variant in product.variants.all():
            values = _spec_map(variant)
            padded_values = [values.get(key, '') for key in keys] + [''] * (max_specs - len(keys))
            variant_row = [
                product.product_code or '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                'true',
                variant.code or '',
                '',
                '',
                *padded_values,
            ]
            row_idx = ws.max_row + 1
            for col_idx, value in enumerate(variant_row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=_cell(value))
                cell.alignment = wrap
                cell.border = border

    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 28
    widths = {
        'A': 18, 'B': 32, 'C': 16, 'D': 18, 'E': 20, 'F': 20, 'G': 20,
        'H': 40, 'I': 22, 'J': 14, 'K': 14, 'L': 12, 'M': 22, 'N': 22,
        'O': 10, 'P': 14, 'Q': 20, 'R': 24, 'S': 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for i in range(len(FIXED_HEADERS) + 1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
